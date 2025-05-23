import os
from PySide6.QtWidgets import QWidget, QVBoxLayout, QLabel, QProgressBar,QHBoxLayout, QFileDialog, QFrame, QMessageBox, QTableWidget, QTableWidgetItem, QSizePolicy                
from PySide6.QtCore import Qt, Signal, QThread
from PySide6.QtGui import QFont, QPixmap
from ui.componentes import BotaoPrimario, BotaoSecundario
from utils.extrator import extrair_dados_fgts_pdfplumber, processar_pasta
from utils.gerador_planilha import gerar_planilha_fgts
from utils.mensagem import mensagem_error, mensagem_sucesso, mensagem_aviso
from utils.icone import usar_icone, recurso_caminho

class WorkerThread(QThread):
    progress = Signal(int)
    progress_detail = Signal(str, int, int)
    finished = Signal(dict, str)
    error = Signal(str)

    def __init__(self, caminho, is_pasta=False):
        super().__init__()
        self.caminho = caminho
        self.is_pasta = is_pasta

    def run(self):
        try:
            if self.is_pasta:
                def progress_callback(arquivo_idx, total_arquivos, arquivo=None, 
                                     pagina=None, total_paginas=None, porcentagem=None):
                    if porcentagem is not None:
                        self.progress.emit(porcentagem)
                    if arquivo and pagina and total_paginas:
                        self.progress_detail.emit(arquivo, pagina, total_paginas)
                
                registros = processar_pasta(self.caminho, progress_callback)
                self.progress.emit(100)
                self.finished.emit(registros, os.path.basename(self.caminho))
            else:
                def progress_callback(pagina_atual, total_paginas):
                    porcentagem = int((pagina_atual / total_paginas) * 100)
                    self.progress.emit(porcentagem)
                    self.progress_detail.emit(os.path.basename(self.caminho), 
                                             pagina_atual, total_paginas)
                
                registros = extrair_dados_fgts_pdfplumber(self.caminho, progress_callback)
                self.progress.emit(100)
                self.finished.emit(registros, os.path.basename(self.caminho))
                
        except Exception as e:
            self.error.emit(str(e))

class TelaExtracao(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Extração de Dados FGTS - Assertivus Contábil")
        self.setGeometry(300, 100, 1200, 720)
        self.setStyleSheet("background-color: #181818; color: #ECECEC; font-family: 'Segoe UI';")
        self.dados_extraidos = {}
        self.nome_arquivo = ""
        self.init_ui()
        usar_icone(self)

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(20)

        header = QFrame()
        header.setFixedHeight(60)
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)
        self.botao_voltar = BotaoSecundario("Voltar ao Dashboard", recurso_caminho("images/voltar.png"))
        self.botao_voltar.clicked.connect(self.voltar_dashboard)
        header_layout.addWidget(self.botao_voltar)
        header_layout.addStretch()
        main_layout.addWidget(header)

        logo_label = QLabel()
        pix = QPixmap(recurso_caminho("images/logo.png"))
        if not pix.isNull():
            logo_label.setPixmap(pix.scaled(180, 180, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo_label.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(logo_label)

        titulo = QLabel("Conversor de PDF FGTS")
        titulo.setFont(QFont("Segoe UI", 24, QFont.Bold))
        titulo.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(titulo)

        botoes_layout = QHBoxLayout()
        botoes_layout.setSpacing(20)
        botoes_layout.setAlignment(Qt.AlignCenter)

        self.botao_selecionar_pdf = BotaoPrimario("Selecionar PDF Único", "#E53935", "#C62828", recurso_caminho("images/pdf.png"))
        self.botao_selecionar_pdf.clicked.connect(lambda: self.selecionar_arquivo(False))
        
        self.botao_selecionar_pasta = BotaoPrimario("Selecionar Pasta de PDFs", "#43A047", "#2E7D32", recurso_caminho("images/pasta.png"))
        self.botao_selecionar_pasta.clicked.connect(lambda: self.selecionar_arquivo(True))

        self.botao_gerar = BotaoPrimario("Gerar Planilha", "#2196F3", "#1976D2", None)
        self.botao_gerar.setEnabled(False)
        self.botao_gerar.clicked.connect(self.gerar_planilha)

        botoes_layout.addWidget(self.botao_selecionar_pdf)
        botoes_layout.addWidget(self.botao_selecionar_pasta)
        botoes_layout.addWidget(self.botao_gerar)

        main_layout.addLayout(botoes_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(400)
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar, alignment=Qt.AlignCenter)

        self.table = QTableWidget()
        self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.table.setVisible(False)
        main_layout.addWidget(self.table)

        self.setLayout(main_layout)

    def voltar_dashboard(self):
        from ui.dashboard import Dashboard
        self.dashboard = Dashboard()
        self.dashboard.showMaximized()
        self.close()

    def selecionar_arquivo(self, is_pasta):
        if is_pasta:
            caminho = QFileDialog.getExistingDirectory(self, "Selecionar Pasta de PDFs", "")
        else:
            caminho = QFileDialog.getOpenFileName(self, "Selecionar PDF", "", "Arquivos PDF (*.pdf)")[0]

        if caminho:
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            
            if not hasattr(self, 'progress_label'):
                self.progress_label = QLabel("Preparando...")
                self.progress_label.setAlignment(Qt.AlignCenter)
                layout_idx = self.layout().indexOf(self.progress_bar)
                self.layout().insertWidget(layout_idx + 1, self.progress_label)
            self.progress_label.setVisible(True)
            
            self.botao_selecionar_pdf.setEnabled(False)
            self.botao_selecionar_pasta.setEnabled(False)

            self.worker = WorkerThread(caminho, is_pasta)
            self.worker.progress.connect(self.atualizar_progresso)
            self.worker.progress_detail.connect(self.atualizar_detalhes_progresso)
            self.worker.finished.connect(self.processamento_concluido)
            self.worker.error.connect(self.erro_processamento)
            self.worker.start()

    def atualizar_progresso(self, valor):
        self.progress_bar.setValue(valor)
    
    def atualizar_detalhes_progresso(self, arquivo, pagina, total_paginas):
        if hasattr(self, 'progress_label'):
            self.progress_label.setText(f"Processando {arquivo} - Página {pagina}/{total_paginas}")

    def processamento_concluido(self, dados, nome_arquivo):
        self.dados_extraidos = dados
        self.nome_arquivo = nome_arquivo
        self.progress_bar.setVisible(False)
        if hasattr(self, 'progress_label'):
            self.progress_label.setVisible(False)
        self.botao_selecionar_pdf.setEnabled(True)
        self.botao_selecionar_pasta.setEnabled(True)

        total_registros = sum(len(registros) for registros in dados.values())
        
        if total_registros == 0:
            QMessageBox.warning(self, "Aviso", "Nenhum dado foi extraído do arquivo.")
            return

        primeira_competencia = list(dados.keys())[0]
        registros = dados[primeira_competencia]
        
        header = ["Matrícula", "Empregado", "CPF", "Admissão", "Base FGTS", "Valor FGTS"]
        self.table.clear()
        self.table.setColumnCount(len(header))
        self.table.setRowCount(len(registros))
        self.table.setHorizontalHeaderLabels(header)

        for r, registro in enumerate(registros):
            self.table.setItem(r, 0, QTableWidgetItem(registro["Matricula"]))
            self.table.setItem(r, 1, QTableWidgetItem(registro["Empregado"]))
            self.table.setItem(r, 2, QTableWidgetItem(registro["CPF"]))
            self.table.setItem(r, 3, QTableWidgetItem(registro["Admissao"]))
            self.table.setItem(r, 4, QTableWidgetItem(registro["Base FGTS"]))
            self.table.setItem(r, 5, QTableWidgetItem(registro["Valor FGTS"]))

        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.resizeRowsToContents()
        self.table.setVisible(True)
        self.botao_gerar.setEnabled(True)
        
        mensagem_sucesso(f"Processamento concluído!\nForam extraídos {total_registros} registros em {len(dados)} competência(s).")

    def erro_processamento(self, mensagem_erro):
        self.progress_bar.setVisible(False)
        self.botao_selecionar_pdf.setEnabled(True)
        self.botao_selecionar_pasta.setEnabled(True)
        mensagem_error(f"Erro ao processar o(s) arquivo(s): {mensagem_erro}")

    def gerar_planilha(self):
        if not self.dados_extraidos:
            mensagem_aviso("Não há dados para gerar a planilha.")
            return

        caminho_saida, _ = QFileDialog.getSaveFileName(self, "Salvar Planilha", "planilha_fgts.xlsx", "Planilhas Excel (*.xlsx)")
        if caminho_saida:
            try:
                import pandas as pd
                from openpyxl import load_workbook
                from openpyxl.styles import Font
                
                with pd.ExcelWriter(caminho_saida, engine='openpyxl') as writer:
                    competencias_ordenadas = sorted(
                        self.dados_extraidos.keys(),
                        key=lambda x: pd.to_datetime(f"01/{x}", dayfirst=True)
                    )

                    for competencia in competencias_ordenadas:
                        registros = self.dados_extraidos[competencia]
                        dados_formatados = []
                        for item in registros:
                            item_copia = item.copy()
                            try:
                                item_copia["Base FGTS"] = f"{float(item_copia['Base FGTS']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                item_copia["Valor FGTS"] = f"{float(item_copia['Valor FGTS']):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                            except (ValueError, TypeError):
                                item_copia["Base FGTS"] = item_copia.get("Base FGTS", "")
                                item_copia["Valor FGTS"] = item_copia.get("Valor FGTS", "")
                            dados_formatados.append(item_copia)

                        df = pd.DataFrame(dados_formatados, columns=["Matricula", "Empregado", "CPF", "Admissao", "Base FGTS", "Valor FGTS"])
                        aba_nome = competencia.replace("/", "_")
                        df.to_excel(writer, sheet_name=aba_nome, index=False)

                wb = load_workbook(caminho_saida)
                fonte_arial = Font(name='Arial', size=10)
                for aba in wb.worksheets:
                    for row in aba.iter_rows(min_row=2):
                        for i, cell in enumerate(row):
                            cell.font = fonte_arial
                            if i == 0:
                                cell.number_format = "@"
                            if i in (4, 5):
                                cell.number_format = '#.##0,00'
                            else:
                                cell.number_format = "@"
                wb.save(caminho_saida)
                
                mensagem_sucesso(f"Planilha gerada com sucesso em:\n{caminho_saida}")
                
                if QMessageBox.question(self, "Planilha salva", "Deseja abrir a planilha agora?", 
                                       QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    import webbrowser
                    webbrowser.open(f"file:///{os.path.abspath(caminho_saida)}")
                    
            except Exception as e:
                mensagem_error(f"Erro ao gerar planilha:\n{str(e)}")